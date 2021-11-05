from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import shutil
import os

#Read the function
def get_input():
    print("Ingrese la funcion con el siguiente formato: 0,1,2,3")
    input_values = input("Funcion: ")
    return [int(value) for value in input_values.split(",")]
 
#Find the number of variables
def get_number_of_variables(max_value):
    num_variables = 2
    indice = 4
    while indice < max_value:
        indice *= 2
        num_variables += 1
    print("Numero de variables: " + str(num_variables))
    return num_variables

#Pass the number from decimal to binary
def to_binary(number, variable_number):
    residuos = []
    isFinished = False
    binary_number = ""
    while not isFinished :
        if(number == 1):
            residuos.append(1)
            isFinished = True
        else:
            residuo = int(number % 2)
            residuos.append(residuo)
            number = int((number - residuo)/2) 
    for part in residuos[::-1]:
        binary_number += str(part)
    while len(binary_number) < variable_number:
        binary_number = "0" + binary_number
    return binary_number

def generate_excel(binary_numbers, variable_number):
    #Excel paths
    base_path = os.getcwd() + "\\mapa_karnaugh\\"
    excel_path = base_path + "excels_base\\" + str(variable_number) + "_var.xlsx"
    new_excel_path = base_path + "excels_finales\\" + str(variable_number) + "_final.xlsx"
    #Creating the new excel
    shutil.copyfile(excel_path, new_excel_path)
    workbook = load_workbook(filename=new_excel_path)
    sheet = workbook.active 
    #Sorting the excel cells
    for row in sheet.rows:
        for cell in row:
            value = str(cell.value)
            if (value != None) and value in binary_numbers:
                cell.fill = PatternFill(start_color='FFE599', end_color='FFE599', fill_type = "solid")
    workbook.save(filename=new_excel_path)
                  
def main():
    function_values = get_input()
    number_of_variables = get_number_of_variables(max(function_values))
    binary_numbers = [] 
    for number in function_values:
        binary_numbers.append(to_binary(number, number_of_variables))
    print("Numeros de la funcion en binario")
    for number in binary_numbers:
        print(number)
    generate_excel(binary_numbers, number_of_variables)
    print("Excel generated")
  
if __name__ == "__main__":
    main()